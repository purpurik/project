import matplotlib.pyplot as plt
import json
import pandas as pd
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QInputDialog,QTableWidgetItem
from tkinter import filedialog


class RatingDialog(QtWidgets.QDialog, QtWidgets.QWidget):
    def __init__(self, students_data):
        super().__init__()
        self.students_data = students_data
        self.setupUi(self)

    def setupUi(self, dialog):
        dialog.setObjectName("Рейтинг")
        dialog.resize(400, 300)
        dialog.setFixedSize(400, 300)
        self.tableWidget = QtWidgets.QTableWidget(dialog)
        self.tableWidget.setGeometry(QtCore.QRect(10, 20, 381, 231))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(2)
        self.tableWidget.setHorizontalHeaderLabels(["Рейтинг", "Средний балл"])
        self.tableWidget.setRowCount(len(self.students_data))
        self.tableWidget.horizontalHeader().setVisible(True)
        self.tableWidget.horizontalHeader().setCascadingSectionResizes(False)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(80)
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.verticalHeader().setVisible(False)
        self.populate_table()

    def populate_table(self):
        sorted_students_data = sorted(self.students_data, key=lambda x: x["average_grade"], reverse=True)
        self.tableWidget.setRowCount(len(sorted_students_data))

        for row in range(len(sorted_students_data)):
            student = sorted_students_data[row]
            name_item = QtWidgets.QTableWidgetItem(student["name"])
            grade_item = QtWidgets.QTableWidgetItem(str(student["average_grade"]))

            self.tableWidget.setItem(row, 0, name_item)
            self.tableWidget.setItem(row, 1, grade_item)


class Ui_valuer(object):
    # Массивы хранящие данные о студенте(ФИО, оценки, средний балл, итоговые оценки, предметы)
    students_data = []
    subjects = []

    def setupUi(self, valuer):
        valuer.setObjectName("valuer")
        valuer.resize(858, 538)
        valuer.setFixedSize(858, 538)
        valuer.setWindowIcon(QtGui.QIcon('student.png'))
        self.centralwidget = QtWidgets.QWidget(valuer)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(10, 20, 713, 421))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(4)
        self.tableWidget.setRowCount(0)
        self.tableWidget.horizontalHeader().setVisible(True)
        self.tableWidget.horizontalHeader().setCascadingSectionResizes(False)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(115)
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.verticalHeader().setVisible(False)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(9)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(9)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(9)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setPointSize(9)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(3, item)
        #studentButton - добавление ячейки
        self.studentButton = QtWidgets.QPushButton(self.centralwidget)
        self.studentButton.setGeometry(QtCore.QRect(732, 18, 111, 56))
        font = QtGui.QFont()
        font.setPointSize(20)
        self.studentButton.setFont(font)
        self.studentButton.setStyleSheet("color: rgb(0, 181, 12);")
        self.studentButton.setIconSize(QtCore.QSize(24, 24))
        self.studentButton.setObjectName("studentButton")
        self.studentButton.clicked.connect(self.createInputDialogStudent)
        #scoreButton - добавление оценок
        self.scoreButton = QtWidgets.QPushButton(self.centralwidget)
        self.scoreButton.setGeometry(QtCore.QRect(490, 450, 233, 71))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.scoreButton.setFont(font)
        self.scoreButton.setObjectName("scoreButton")
        self.scoreButton.clicked.connect(self.createInputDialogGrade)
        #removeButton - удаление ячеек
        self.removeButton = QtWidgets.QPushButton(self.centralwidget)
        self.removeButton.setGeometry(QtCore.QRect(732, 360, 111, 39))
        self.removeButton.setText("Удалить студента")
        self.removeButton.setStyleSheet("color: rgb(200, 50, 30);")
        self.removeButton.setIconSize(QtCore.QSize(24, 24))
        self.removeButton.setObjectName("removeButton")
        self.removeButton.clicked.connect(self.deleteRow)
        #removeAllButton - удаление ячеек
        self.removeAllButton = QtWidgets.QPushButton(self.centralwidget)
        self.removeAllButton.setGeometry(QtCore.QRect(732, 450, 111, 71))
        self.removeAllButton.setText("")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("icons/delete.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.removeAllButton.setIcon(icon)
        self.removeAllButton.setIconSize(QtCore.QSize(24, 24))
        self.removeAllButton.setStyleSheet("color: rgb(200, 50, 30);")
        self.removeAllButton.setObjectName("removeButton")
        self.removeAllButton.clicked.connect(self.deleteAll)
        #removeLessonButton - удаление предмета
        self.removeLessonButton = QtWidgets.QPushButton(self.centralwidget)
        self.removeLessonButton.setGeometry(QtCore.QRect(732, 403, 111, 39))
        self.removeLessonButton.setText("Удалить предмет")
        self.removeLessonButton.setStyleSheet("color: rgb(200, 50, 30);")
        self.removeLessonButton.setIconSize(QtCore.QSize(24, 24))
        self.removeLessonButton.setObjectName("removeButton")
        self.removeLessonButton.clicked.connect(self.deleteLesson)
        #nameButton - добавление имени
        self.nameButton = QtWidgets.QPushButton(self.centralwidget)
        self.nameButton.setGeometry(QtCore.QRect(10, 450, 233, 71))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.nameButton.setFont(font)
        self.nameButton.setObjectName("nameButton")
        self.nameButton.clicked.connect(self.createInputDialogName)
        #lessonButton - добавление предмета
        self.lessonButton = QtWidgets.QPushButton(self.centralwidget)
        self.lessonButton.setGeometry(QtCore.QRect(250, 450, 233, 71))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.lessonButton.setFont(font)
        self.lessonButton.setObjectName("lessonButton")
        self.lessonButton.clicked.connect(self.createInputDialogLesson)
        #ratingButton - рейтинг
        self.ratingButton = QtWidgets.QPushButton(self.centralwidget)
        self.ratingButton.setGeometry(QtCore.QRect(732, 81, 111, 50))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.ratingButton.setFont(font)
        self.ratingButton.setStyleSheet("color: rgb(248, 248, 9);")
        self.ratingButton.setText("")
        self.ratingButton.clicked.connect(self.show_rating_dialog)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("icons/star.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.ratingButton.setIcon(icon1)
        self.ratingButton.setIconSize(QtCore.QSize(24, 24))
        self.ratingButton.setObjectName("ratingButton")
        #diagramButton - диаграмма
        self.diagramButton = QtWidgets.QPushButton(self.centralwidget)
        self.diagramButton.setGeometry(QtCore.QRect(732, 136, 111, 53))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.diagramButton.setFont(font)
        self.diagramButton.setStyleSheet("color: rgb(248, 248, 9);")
        self.diagramButton.setText("")
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("icons/circular-diagram.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.diagramButton.setIcon(icon2)
        self.diagramButton.setIconSize(QtCore.QSize(32, 32))
        self.diagramButton.setObjectName("diagramButton")
        self.diagramButton.clicked.connect(self.showDiagram)
        # self.excelImportButton = QtWidgets.QPushButton(self.centralwidget)
        # self.excelImportButton.setGeometry(QtCore.QRect(732, 450, 111, 35))
        # font = QtGui.QFont()
    
        # font.setPointSize(10)
        # self.excelImportButton.setFont(font)
        # self.excelImportButton.setStyleSheet("color: rgb(0, 181, 12);")
        # self.excelImportButton.setText("IMPORT")
        # icon2 = QtGui.QIcon()
        # icon2.addPixmap(QtGui.QPixmap("icons/excel.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        # self.excelImportButton.setIcon(icon2)
        # self.excelImportButton.setIconSize(QtCore.QSize(32, 32))
        # self.excelImportButton.setObjectName("diagramButton")
        # self.excelImportButton.clicked.connect(self.load_data_excel)

        #excelSaveButton - сохранение таблицы в excel
        self.excelSaveButton = QtWidgets.QPushButton(self.centralwidget)
        self.excelSaveButton.setGeometry(QtCore.QRect(732, 194, 111, 50))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.excelSaveButton.setFont(font)
        self.excelSaveButton.setStyleSheet("color: rgb(0, 181, 12);")
        self.excelSaveButton.setIconSize(QtCore.QSize(32, 32))
        self.excelSaveButton.setObjectName("excelButton")
        self.excelSaveButton.setText("Сохранить Excel")
        self.excelSaveButton.clicked.connect(self.save_excel)
        #saveButton - сохранение таблицы в json
        self.saveButton = QtWidgets.QPushButton(self.centralwidget)
        self.saveButton.setGeometry(QtCore.QRect(732, 249, 111, 50))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.saveButton.setFont(font)
        self.saveButton.setStyleSheet("color: rgb(103, 118, 250);")
        self.saveButton.setIconSize(QtCore.QSize(24, 24))
        self.saveButton.setObjectName("saveButton")
        self.saveButton.clicked.connect(self.save_data)
        #importButton - импорт таблицы с json
        self.importButton = QtWidgets.QPushButton(self.centralwidget)
        self.importButton.setGeometry(QtCore.QRect(732, 304, 111, 50))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.importButton.setFont(font)
        self.importButton.setStyleSheet("color: rgb(103, 118, 250);")
        self.importButton.setIconSize(QtCore.QSize(24, 24))
        self.importButton.setObjectName("importButton")
        self.importButton.clicked.connect(self.load_data)

        valuer.setCentralWidget(self.centralwidget)

        self.retranslateUi(valuer)
        QtCore.QMetaObject.connectSlotsByName(valuer)

    def retranslateUi(self, valuer):
        _translate = QtCore.QCoreApplication.translate
        valuer.setWindowTitle(_translate("valuer", "Valuer"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("valuer", "ФИО"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("valuer", "Предмет"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("valuer", "Средний балл"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("valuer", "Оценка"))
        self.studentButton.setText(_translate("valuer", "+"))
        self.scoreButton.setText(_translate("valuer", "Добавить оценки"))
        self.nameButton.setText(_translate("valuer", "Добавить имя"))
        self.lessonButton.setText(_translate("valuer", "Добавить предмет"))
        self.saveButton.setText(_translate("valuer", "Сохранить json"))
        self.importButton.setText(_translate("valuer", "Импорт в json"))

    # Функция сохраняет данные в файл формата json
    def save_data(self):
        if self.students_data and self.subjects:
            file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
            try:
                with open(file_path, 'w') as f:
                    json.dump({'students_data': self.students_data, 'subjects': self.subjects}, f)
            except FileNotFoundError:
                print('Файл не найден')
            
        else:
            print("Нету данных для сохранения!")

    def save_excel(self):
        if self.students_data:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            try:
                data_renamed = []
                for student_data in self.students_data:
                    student_data_renamed = {
                        'Имя': student_data.pop('name'),
                    }
                    for i, subject in enumerate(self.subjects, 1):
                        if f'lesson{i}' in student_data:
                            student_data_renamed[subject] = student_data.pop(f'lesson{i}')
                    if 'average_grade' in student_data:
                        student_data_renamed['Средний балл'] = student_data.pop('average_grade')
                    if 'grade' in student_data:
                        student_data_renamed['Оценка'] = student_data.pop('grade')

                    student_data_renamed.update({k: v for k, v in student_data.items() if k not in ['name', 'grades', 'average_grade', 'grade']})
                    data_renamed.append(student_data_renamed)

                df = pd.DataFrame(data_renamed)
                df.to_excel(file_path, index=False)

                column_widths = {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20, 'G': 20, 'H': 20, 'I': 20, 'J': 20}
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active
                for column, width in column_widths.items():
                    worksheet.column_dimensions[column].width = width
                workbook.save(file_path)

                print("Данные сохранены удачно.")
            except FileNotFoundError:
                print('FileNotFoundError: Некорректный путь к файлу')
            except KeyError:
                print('KeyError: Ключ не найден')
            except ValueError:
                print('ValuerError: Отмена выбора')
        else:
            print("No data to save!")

    # Загрузка данных с файла
    def load_data(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
            if not file_path:
                return
            with open(file_path) as f:
                data = json.load(f)
                if self.students_data == data['students_data'] and self.subjects == data['subjects']:
                    print('Ты уже загрузил данные!')
                else:
                    self.tableWidget.setRowCount(0)
                    new_row_position = self.tableWidget.rowCount()
                    self.students_data = data['students_data']
                    self.subjects = data['subjects']
                    self.updateTableHeaders()
                    for i in self.students_data:
                        self.tableWidget.insertRow(new_row_position)
                    self.data()
                    self.calculate()
                    print("Данные загружены - Удачно.")

        except FileNotFoundError as ex:
            print('FileNotFoundError: Некорректный путь к файлу')
            print(ex)
            self.students_data = []
            self.subjects = []
        except KeyError as ex:
            print('KeyError: В файле нету данных')
            self.students_data = []
            self.subjects = []

    def load_data_excel(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return
            df = pd.read_excel(file_path)

            # Convert the DataFrame to a list of dictionaries
            data = df.to_dict(orient='records')

            # # Clear the current data and subjects lists
            # self.students_data = []
            # self.subjects = []

            # Add the data to the students_data list
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            self.tableWidget.setRowCount(sheet.max_row + 1)
            self.tableWidget.setRowCount(sheet.max_column)

            list_values = list(sheet.values)
            self.tableWidget.setHorizontalHeaderLabels(list_values[0])
            
            row_index = 0
            for value_tuple in list_values[1:]:
                column_index = 0
                for value in value_tuple:
                    self.tableWidget.setItem(row_index ,column_index ,QTableWidgetItem(str(value)))
                    column_index += 1
                row_index += 1


            # Update the table headers and display the data
            # self.updateTableHeaders()
            # self.data()
            # self.calculate()
            print("Data loaded - Successfully.")

        except FileNotFoundError as ex:
            print('FileNotFoundError: Некорректный путь к файлу')
            print(ex)
            self.students_data = []
            self.subjects = []
        except Exception as ex:
            print('Exception:', ex)
            self.students_data = []
            self.subjects = []

    # Функция добавляет предмет в массив subjects
    def createInputDialogLesson(self):
        lesson, ok = QInputDialog.getText(None, 'Добавить предмет', 'Введите название предмета')
        if ok and lesson:
            for row in range(self.tableWidget.rowCount()):
                self.tableWidget.setItem(row, len(self.subjects) + 1, QtWidgets.QTableWidgetItem())
            self.subjects.append(lesson)
            print(self.subjects)
            self.updateTableHeaders()
            self.calculate()

    # Функция добавляет предметы из массива subjects в колонны 
    def updateTableHeaders(self):
        header_labels = ["ФИО"] + self.subjects + ["Средний балл", "Оценка"]
        self.tableWidget.setColumnCount(len(header_labels))
        self.tableWidget.setHorizontalHeaderLabels(header_labels)

    # Функция распределяет данные имени, среднего балла и оценок по нужным колоннам
    def data(self):
        row = 0
        for student in self.students_data:
            self.tableWidget.setItem(row, 0, QTableWidgetItem(student['name']))
            self.tableWidget.setItem(row, len(self.subjects) + 1, QTableWidgetItem(str(round(student['average_grade']))))
            self.tableWidget.setItem(row, len(self.subjects) + 2, QTableWidgetItem(str(student['grade'])))
            for i in range(1, len(self.subjects) + 1):
                self.tableWidget.setItem(row, i, QTableWidgetItem(str(student['lesson{}'.format(i)])))
            row += 1


    def showDiagram(self):
        grades = [student['average_grade'] for student in self.students_data]
        names = [student['name'] for student in self.students_data]

        plt.figure(figsize=(10, 6))
        plt.bar(names, grades, color='skyblue')
        plt.xlabel('Студенты')
        plt.ylabel('Средний балл')
        plt.title('Диаграмма баллов студентов')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.show()
        pass


    def createInputDialogGrade(self):
        print(self.students_data,'\n', self.subjects)
        if self.tableWidget.rowCount() == 0:
            QtWidgets.QMessageBox.warning(
                    self.centralwidget,
                    "!ОШИБКА!",
                    "ДОБАВЬ СТУДЕНТА!",
                    QtWidgets.QMessageBox.Ok
                )
        elif self.subjects == []:
            QtWidgets.QMessageBox.warning(
                    self.centralwidget,
                    "!ОШИБКА!",
                    "ДОБАВЬ ПРЕДМЕТ!",
                    QtWidgets.QMessageBox.Ok
                )
        else:
            count, ok = QInputDialog.getText(None, 'graduate.', 'Введите кол-во оценок')
            try:
                if ok and count:
                    lesson_grade = []
                    selected_row = self.tableWidget.currentRow()
                    student_data = self.students_data[selected_row]
                    selected_column = self.tableWidget.currentColumn()
                    lesson_key = 'lesson{}'.format(selected_column) 
                    lesson_grades = [] if 'lesson{}'.format(selected_column) not in student_data else student_data['lesson{}'.format(selected_column)]
                    student_data[lesson_key] = lesson_grade
                    if selected_row < 0 or selected_row >= len(self.students_data):
                        return

                    grades = [] if 'grades' not in student_data else student_data['grades']

                    print(len(lesson_grades))
                    for l in lesson_grades:
                        for i in student_data['grades']:
                                for j in lesson_grades:
                                    if j == i:
                                        student_data['grades'].remove(i and j)

                    for i in range(int(count)):
                        text, ok = QInputDialog.getText(None, 'graduate.', 'Введите оценку')
                        if ok:
                            entered_grade = int(text)
                            grades.append(entered_grade)
                            lesson_grade.append(entered_grade)
                    self.tableWidget.setItem(selected_row, selected_column, QtWidgets.QTableWidgetItem(str(lesson_grade)))
                    self.calculate()

            except ValueError:
                print('write the number')

            except TypeError as e:
                print(e)
                print('Вы ввели некорекктное значение')


    def calculate(self):
        if not self.students_data:
            print('No students data to calculate.')
            return

        for row in range(self.tableWidget.rowCount()):
            student_data = self.students_data[row]
            grades = student_data.get('grades', [])

            if not grades:
                average_grade = 0
            else:
                average_grade = sum(grades) / len(grades)

            student_data['average_grade'] = average_grade

            if average_grade >= 90:
                grade = '5'
            elif average_grade >= 70:
                grade = '4'
            elif average_grade >= 60:
                grade = '3'
            else:
                grade = '2'

            student_data['grade'] = grade

            self.tableWidget.setItem(row, len(self.subjects) + 1, QtWidgets.QTableWidgetItem(str(round(average_grade, 2))))
            self.tableWidget.setItem(row, len(self.subjects) + 2, QtWidgets.QTableWidgetItem(str(grade)))


    def createInputDialogStudent(self):
        self.students_data.append({"name": '', "grades": [], "average_grade": 0, "grade":0})
        new_row_position = self.tableWidget.rowCount()
        self.tableWidget.insertRow(new_row_position)

    
    def createInputDialogName(self):
        try:
            selected_row = self.tableWidget.currentRow()
            if self.tableWidget.rowCount() == 0:
                QtWidgets.QMessageBox.warning(
                        self.centralwidget,
                        "!ОШИБКА!",
                        "ДОБАВЬ СТУДЕНТА!",
                        QtWidgets.QMessageBox.Ok
                    )
            
            elif selected_row < 0 or selected_row >= len(self.students_data):
                return

            student_data = self.students_data[selected_row]
            name, ok = QInputDialog.getText(None, 'graduate.', 'Введите имя студента')
            if ok:
                student_data['name'] = name

            self.tableWidget.setItem(selected_row, 0, QtWidgets.QTableWidgetItem(str(name)))
        except IndexError:
            print('indexerror')


    def deleteLesson(self):
        row = 0
        selected_column = self.tableWidget.currentColumn()
        try:
            current_lesson = str('lesson{}'.format(selected_column))
            del self.subjects[selected_column - 1]
            self.tableWidget.removeColumn(selected_column)
            print(self.subjects, ' - deleteLesson')
            for student in self.students_data:
                student.pop(current_lesson)
            self.calculate()
        except KeyError as ex:
            print(ex)

    def deleteRow(self):
        if self.tableWidget.rowCount() == 0:
             QtWidgets.QMessageBox.warning(
                    self.centralwidget,
                    "!ОШИБКА!",
                    "ДОБАВЬ СТУДЕНТА!",
                    QtWidgets.QMessageBox.Ok
                )
        else:
            selected_row = self.tableWidget.currentRow()
            if selected_row >= 0:
                    self.tableWidget.removeRow(selected_row)
                    self.students_data.pop(selected_row)

    def deleteAll(self):
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.subjects = []
        self.students_data = []
        self.updateTableHeaders()
        self.data()
        self.calculate()

    def show_rating_dialog(self):
        dialog = RatingDialog(self.students_data)
        dialog.exec_()

if __name__ == "__main__":
    import sys 
    app = QtWidgets.QApplication(sys.argv) 
    valuer = QtWidgets.QMainWindow() 
    ui = Ui_valuer() 
    ui.setupUi(valuer) 
    valuer.show() 
    sys.exit(app.exec_())